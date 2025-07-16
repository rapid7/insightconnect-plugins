import base64
import io
import logging
import urllib.parse
import zipfile
from datetime import datetime
from difflib import get_close_matches
from typing import Any, Dict, List, Union

import openpyxl
import pdfplumber
import regex
import tldextract
import validators
from insightconnect_plugin_runtime.exceptions import PluginException
from odf.opendocument import load
from openpyxl.workbook.workbook import Worksheet
from pdfminer.pdfparser import PDFSyntaxError
from pdfplumber.page import Page
from publicsuffix2 import PublicSuffixList

from icon_extractit.util.util import DateFormatStrings, Regex

DEFAULT_ENCODING = "utf-8"
DEFAULT_PDF_WRAPPING_TOLERANCE = 5.0


def _get_cell_number_format(cell_number_format: str) -> str:
    """The function reformats obtained cell number format, formatted for datetime strftime() method and returns it

    :param cell_number_format: Cell number format obtained from openpyxl
    :type cell_number_format: str

    :return: Returns formatted for datetime cell number format
    :rtype: str
    """

    return (
        cell_number_format.replace("mm", "%m")
        .replace("dd", "%d")
        .replace("yyyy", "%Y")
        .replace("yy", "%y")
        .replace('"', "")
        .lstrip("0")
    )


def _extract_dates_from_cells(input_sheet: Worksheet) -> List[str]:
    """Extracts the dates from all non empty cells that are formatted as Date and Time cells
    with the date format defined in the xlsx sheet

    :param input_sheet: Object of the worksheet from where the data will be obtained
    :type input_sheet: Worksheet

    :return: Returns list containing Date Time values from the input cells
    :rtype: List[str]
    """

    output_dates = []
    for row in input_sheet.iter_rows():
        for cell in row:
            if cell.value and isinstance(cell.value, datetime):
                strf_date = cell.value.strftime(_get_cell_number_format(cell.number_format))
                strf_date = strf_date.replace("m", "%m") if strf_date.count("m") == 1 else strf_date
                strf_date = strf_date.replace("d", "%d") if strf_date.count("d") == 1 else strf_date
                strf_date = cell.value.strftime(strf_date)
                output_dates.append(strf_date)
    return output_dates


def _extract_dates_from_xlsx(input_file: Any) -> str:
    """Extracts all the dates from .xlsx file

    :param input_file: Opened in buffer xlsx file
    :type input_file: Any

    :return: Str containing all the extracted date and time values separated by space character
    :rtype: str
    """

    try:
        workbook = openpyxl.load_workbook(input_file)
        sheet_names = workbook.sheetnames
        output_list_of_dates = []
        for sheet_name in sheet_names:
            sheet = workbook[sheet_name]
            output_list_of_dates += _extract_dates_from_cells(sheet)
        return " ".join(output_list_of_dates)
    except Exception:
        return ""


def extract(
    provided_regex: str,
    provided_string: str,
    provided_file: str,
    keep_original_url: bool = False,
) -> list:
    matches = []
    if provided_string:
        if keep_original_url:
            matches = regex.findall(provided_regex, provided_string)  # regex finds both encoded and unencoded urls
        else:
            provided_string = urllib.parse.unquote(provided_string)
            matches = regex.findall(provided_regex, provided_string)
    elif provided_file:
        try:
            if keep_original_url:
                provided_file = base64.b64decode(provided_file.encode(DEFAULT_ENCODING)).decode(DEFAULT_ENCODING)
                matches = regex.findall(provided_regex, provided_file)
            else:
                provided_file = urllib.parse.unquote(
                    base64.b64decode(provided_file.encode(DEFAULT_ENCODING)).decode(DEFAULT_ENCODING)
                )
                matches = regex.findall(provided_regex, provided_file)
        except UnicodeDecodeError:
            file_content = extract_content_from_file(base64.b64decode(provided_file), provided_regex)
            matches = regex.findall(provided_regex, file_content)
    return list(dict.fromkeys(matches))


def extract_all_date_formats(provided_string: str, provided_file: str) -> list:
    matches = []
    if provided_string:
        provided_string = urllib.parse.unquote(provided_string)
        for regex_pattern in Regex.HumanToRegexPatterns.values():
            matches += regex.findall(regex_pattern, provided_string)
            provided_string = regex.sub(regex_pattern, "", provided_string)
    elif provided_file:
        try:
            provided_file = urllib.parse.unquote(
                base64.b64decode(provided_file.encode(DEFAULT_ENCODING)).decode(DEFAULT_ENCODING)
            )
            for regex_pattern in Regex.HumanToRegexPatterns.values():
                matches += regex.findall(regex_pattern, provided_file)
                provided_file = regex.sub(regex_pattern, "", provided_file)
        except UnicodeDecodeError:
            file_content = extract_content_from_file(base64.b64decode(provided_file))
            for regex_pattern in Regex.HumanToRegexPatterns.values():
                matches += regex.findall(regex_pattern, file_content)
                file_content = regex.sub(regex_pattern, "", file_content)
    return list(dict.fromkeys(matches))


def extract_filepath(provided_regex: str, provided_string: str, provided_file: str) -> list:
    matches = []
    if provided_string:
        new_string = regex.sub(Regex.URL, "", provided_string)
        new_string = regex.sub(Regex.Date, "", new_string)
        matches = regex.findall(provided_regex, new_string)
    elif provided_file:
        try:
            new_file = base64.b64decode(provided_file.encode(DEFAULT_ENCODING)).decode(DEFAULT_ENCODING)
        except UnicodeDecodeError:
            new_file = extract_content_from_file(base64.b64decode(provided_file), provided_regex)
        new_file = regex.sub(Regex.URL, "", new_file)
        new_file = regex.sub(Regex.Date, "", new_file)
        matches = regex.findall(provided_regex, new_file)
    return list(dict.fromkeys(matches))


def extract_content_from_file(provided_file: bytes, provided_regex: str = "") -> str:  # noqa: C901, MC0001
    with io.BytesIO(provided_file) as file_:
        try:
            # extracting content from DOCX, PPTX, XLSX, ODT, ODP, ODF files
            with zipfile.ZipFile(file_) as unzip_files:
                files_content = ""
                x = unzip_files.infolist()
                for i in enumerate(x):
                    try:
                        files_content += unzip_files.read(x[i[0]]).decode(DEFAULT_ENCODING)
                    # After unpacking, may contain images for which the decoding will fail
                    except UnicodeDecodeError:
                        continue
                # check if file is xlsx
                xlsx_output = _extract_dates_from_xlsx(file_)

            # Check for ODT, ODP, ODF File content without metadata
            odf_files_content = load_libreoffice_file_content(file_)
            if odf_files_content:
                files_content = odf_files_content

            # remove xml tags from files content
            content_without_xml_tags = regex.sub(
                r"<[\p{L}\p{N}\p{Lo}\p{So} :\/.\"=_%,(){}+#&;?-]*>", " ", files_content
            )
            return " ".join([content_without_xml_tags.replace("\r", ""), xlsx_output])
        except zipfile.BadZipFile:
            try:
                logging.getLogger("pdfminer").setLevel(logging.WARNING)
                # extracting content from PDF file
                pdf_content = ""
                with pdfplumber.open(file_) as pdf_file:
                    for page in pdf_file.pages:
                        page_content = page.extract_text()
                        for word in extract_wrapped_words_from_pdf_page(page, provided_regex):
                            page_content = page_content.replace(word, word.replace("\n", ""))
                        pdf_content += page_content
                return pdf_content
            except PDFSyntaxError:
                raise PluginException(
                    cause="The type of the provided file is not supported.",
                    assistance="Supported file types are text/binary, such as: PDF, DOCX, PPTX, XLSX, ODT, ODP, ODF, TXT, ZIP",
                )


def extract_wrapped_words_from_pdf_page(
    page: Page,
    provided_regex: str = "",
    tolerance: float = DEFAULT_PDF_WRAPPING_TOLERANCE,
) -> List[str]:
    """
    Extract wrapped words from a PDF page.

    :param page: The PDF page from which to extract wrapped words.
    :type: Page

    :param provided_regex: The regex for the type of words to be searched for, e.g. email/domain format.
    :type: str

    :param tolerance: The tolerance value for detecting wrapped words. Defaults to DEFAULT_PDF_WRAPPING_TOLERANCE.
    :type: float

    :return: A list of wrapped words extracted from the PDF page.
    :rtype: List[str]
    """

    wrapped_words = []
    max_x1 = max(character.get("x1") for character in page.chars)
    extracted_words = page.extract_words(use_text_flow=True)

    for index, word in enumerate(extracted_words):
        if (max_x1 - word.get("x1")) < tolerance:
            # if the current or next word in the list are valid matches then do not try and join then
            if provided_regex:
                if (
                    not regex.findall(provided_regex, word.get("text", ""))
                    and (index + 1) < len(extracted_words)
                    and not regex.findall(provided_regex, extracted_words[index + 1].get("text", ""))
                ):
                    wrapped_words.append(f"{word.get('text')}\n")
            else:
                wrapped_words.append(f"{word.get('text')}\n")
    return wrapped_words


def strip_subdomains(matches: list) -> list:
    for match in enumerate(matches):
        stripped_domain = tldextract.extract(match[1])
        # In some cases, tldextract recognizes a suffix as a domain, adds the domain to subdomain, and returns an empty
        # string as suffix, so we check that tldextract extracted the suffix.
        if not stripped_domain.suffix:
            suffix = stripped_domain.domain
            subdomain = stripped_domain.subdomain
            if subdomain and suffix:
                # here we split `subdomain` and extract the domain, which is the last element from the `subdomains` list
                subdomains = subdomain.split(".")
                matches[match[0]] = f"{subdomains[len(subdomains) - 1]}.{suffix}"
        else:
            matches[match[0]] = f"{stripped_domain.domain}.{stripped_domain.suffix}"
    return list(dict.fromkeys(matches))


def clear_domains(matches: list) -> list:
    # The method is designed to reduce the number of false positives results. The domain regex has been extended to
    # include the path from the URL and the @ and = characters if they are at the end of the found domain.
    new_matches = []
    for match in enumerate(matches):
        # Here we remove the path, this solution allows to prevent the situation where the file names from the URL will
        # be extracted as domains, e.g. www.example.com/test.html
        split_match = match[1].split("/")[0]
        # Here we remove the port number, if present, which could prevent a legitimate domain name match.
        # Ex. ssh.example.com:22 becomes ssh.example.com
        split_match = split_match.split(":")[0]
        # Check if the match ends with a common file extension. If there is an invalid suffix it will be ''
        if tldextract.extract(split_match).suffix:
            # Here we eliminate all domains that end with = or @. This avoids extracting two domains from an email address,
            # e.g. user.test@example.com, or extracting field names from an EML file as domains, e.g. two domains would be
            # extracted from "header.from=example.com"
            if not split_match.endswith("@") and not split_match.endswith("="):
                new_matches.append(split_match.lower())
    return list(dict.fromkeys(new_matches))


def remove_extracted_urls_from_links(input_list: list) -> list:
    """
    remove_extracted_urls_from_links This function allows to remove extracted links from URLs
    :param input_list: List of extracted URLs
    :return: List without additionally extracted URLs from link params
    """
    output_list = input_list.copy()
    for url in input_list:
        temporary_list = input_list.copy()
        temp_url = f"={temporary_list.pop(temporary_list.index(url))}"
        for url_check in temporary_list:
            if temp_url.replace("&amp", "") in url_check:
                output_list.pop(output_list.index(temp_url[1:]))
                break
    return output_list


def clear_urls(matches: list) -> list:
    new_matches = []
    for match in enumerate(matches):
        if not validators.ip_address.ipv4(match[1]) and not validators.email(match[1]):
            new_matches.append(match[1])
    return new_matches


def clear_emails(matches: list) -> list:
    new_matches = []
    for match in enumerate(matches):
        if validators.email(match[1]):
            new_matches.append(match[1])
    return new_matches


def define_linux_date_time_format(date_format: str) -> str:
    return DateFormatStrings.human_to_linux_mapping.get(date_format)


def define_date_time_regex(date_format: str) -> str:
    return Regex.HumanToRegexPatterns.get(date_format)


def parse_time_all_date_formats(dates: list) -> list:
    for date in enumerate(dates):
        for linux_date_time_format in DateFormatStrings.human_to_linux_mapping.values():
            date_value_separators = regex.findall(
                r"[^(%b)(%d)(%m)(%y)(%h)(%s)(%D)(%M)(%Y)(%H)(%S)]", linux_date_time_format
            )
            # Alternatively insert date and separators into list that is then joined together as a string. Allows
            # for retrieved dates to be read in an extractable format and account for all future date format variations.
            date_string_list = list(range((len(date[1]) + len(date_value_separators))))
            date_string_list[::2] = date[1]
            date_string_list[1::2] = date_value_separators
            date_string = "".join(date_string_list)
            try:
                date_time_obj = datetime.strptime(date_string, linux_date_time_format)
                dates[date[0]] = date_time_obj.strftime("%Y-%m-%dT%H:%M:%SZ")
                break
            except ValueError:
                continue
    return dates


def parse_time(dates: list, date_format: str) -> list:
    linux_date_time_format = define_linux_date_time_format(date_format)
    # Regex pattern used for identification of separators to handle date time formats in ISO format and all possible
    # cases of mixed separator format that can be converted to a datetime object
    date_value_separators = regex.findall(r"[^(%b)(%d)(%m)(%y)(%h)(%s)(%D)(%M)(%Y)(%H)(%S)]", linux_date_time_format)

    for date in enumerate(dates):
        # Alternatively insert date and separators into list that is then joined together as a string. Allows
        # for retrieved dates to be read in an extractable format and account for all future date format variations.
        date_string_list = list(range((len(date[1]) + len(date_value_separators))))
        date_string_list[::2] = date[1]
        date_string_list[1::2] = date_value_separators
        date_string = "".join(date_string_list)
        try:
            date_time_obj = datetime.strptime(date_string, linux_date_time_format)
            dates[date[0]] = date_time_obj.strftime("%Y-%m-%dT%H:%M:%SZ")
        except ValueError:
            raise PluginException(
                cause=f"The found date {date_string} could not be parsed as a real date",
                assistance="Please review selected date format input for date extraction",
            )
    return dates


def fix_emails_suffix(emails: List[str]) -> List[str]:
    """
    Fixes the suffix of a list of email addresses.

    :param emails: A list of email addresses to be checked and potentially modified.
    :type: List[str]

    :return: A list of email addresses with the correct suffix.
    :rtype: List[str]
    """

    public_suffix_list = PublicSuffixList()
    tlds_set = set(public_suffix_list.tlds)

    emails_copy = emails.copy()
    try:
        for index, email in enumerate(emails_copy):
            extracted_suffix = public_suffix_list.get_tld(email)
            if extracted_suffix in tlds_set:
                continue
            possible_match = get_close_matches(extracted_suffix, tlds_set, n=1, cutoff=0)[0]
            emails_copy[index] = email.replace(extracted_suffix, possible_match)
    except Exception:
        return emails
    return emails_copy


def load_libreoffice_file_content(file_: Union[bytes, io.BytesIO]) -> Union[str, None]:
    try:
        document = load(file_)
        return document.toXml()
    except Exception:
        return None
